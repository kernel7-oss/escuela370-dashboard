import openpyxl
import re
import os

# Existing database mappings from data.js
DOCENTES_MAPPING = {
    "CARLA": "d1",
    "MILTON": "d2",
    "CLAUDIA": "d3",
    "FERNANDO": "d4",
    "CRISTIAN": "d6",
    "YESICA": "d7",
    "GABRIELA": "d8",
    "AMELIA": "d9",
    "CORINA": "d21",
    "MARIA A.": "d22",
    "PAOLA": "d11",
    "ELIANA": "d20",
    "VICTOR": "d29",
    "LAGOS SILVIA": "d30",
    "VENTER RICARDO": "d18",
    "TORANZOS PABLO": "d15",
    "CUENCA ANA": "d17",
    "BULACIOS VANINA": "d14",
    "GALVÁN ESTRELLA": "d16",
    "GONZALEZ KARINA": "d19"
}

STUDENTS_MAPPING = {
    "BENICIO": "e1", "LÓPEZ BENICIO": "e1", "LOPEZ BENICIO": "e1",
    "ASTOR": "e2", "CARDENAS ASTOR": "e2", "CÁRDENAS ASTOR": "e2",
    "MIRKO": "e3", "ESPINA MIRKO": "e3",
    "CHRISTIAN": "e4", "CARRIZO CHRISTIAN": "e4", "CARRIZO CHRISTIAN NÉSTOR": "e4", "CARRIZO CHRISTIAN NESTOR": "e4",
    "LORENZO": "e5", "MONSON LORENZO": "e5", "MONZON LORENZO": "e5",
    "ISABELLA": "e6", "BERLAY ISABELLA": "e6", "BERLAY ISABELLA VICTORIA": "e6",
    "NICOLÁS": "e7", "MUÑOZ G. T. NICOLÁS": "e7", "MUÑOZ G. T. NICOLAS": "e7", "MUÑOZ G.T. NICOLÁS": "e7",
    "VILLAFAÑE": "e8", "VILLAFAÑE POLTI ISABELLA": "e8", "VILLAFAÑE POLTI ISABELA": "e8",
    "PAULO": "e12", "ROMERO PAULO DANIEL": "e12", "ROMERO PAULO": "e12",
    "MURIEL": "e13", "GONZALEZ MAZA MURIEL": "e13", "GONZÁLEZ MAZA MURIEL": "e13",
    "AGUSTÍN": "e14", "VIDAL AGUSTÍN": "e14", "VIDAL AGUSTÍN IGNACIO": "e14", "VIDAL AGUSTIN IGNACIO": "e14",
    "DYLAN": "e15", "ARIAS LEAL DYLAN": "e15", "ARIAS LEAL DYLAN EXEQUIEL": "e15",
    "JERÓNIMO": "e16", "JULIO JERÓNIMO": "e16", "JULIO JERÓNIMO JESÚS": "e16", "JULIO JERONIMO JESUS": "e16",
    "LUCAS": "e17", "CASTRO LUCAS": "e17", "CASTRO LUCAS ALEJANDRO": "e17",
    "CAROLINA": "e18", "PEREZ CAROLINA": "e18", "PEREZ CAROLINA GISELLE": "e18", "PEREZ CAROLINA GISELE": "e18",
    "JONAS": "e19", "CÁRCAMO JONAS": "e19", "CARCAMO JONAS": "e19",
    "MARTINA": "e20", "OLIVERA HERRERA MARTINA FRANCESCA": "e20",
    "ROMÁN": "e21", "GONZALEZ ROMÁN ANTONIO": "e21", "GONZÁLEZ ROMÁN": "e21",
    "RENZO": "e22", "COMINELLI PICON RENZO": "e22", "COMINELLI PICÓN RENZO": "e22",
    "JUAN IGNACIO": "e23", "VERA JUAN IGNACIO": "e23",
    "AMBAR": "e24", "CACERES MARIN AMBAR": "e24", "CACERES MARIN AMBAR SOFIA": "e24", "CÁCERES MARIN AMBAR SOFIA": "e24",
    "SOFIA MORENA": "e25", "RUIZ SOFIA": "e25", "RUIZ SOFIA MORENA": "e25",
    "BENJAMIN URIEL": "e26", "MUÑOZ DEMARIA BENJAMIN URIEL": "e26", "MUÑOZ DEMARIA BENJAMIN": "e26",
    "ENZO": "e27", "RIOS ENZO": "e27", "RIOS ENZO LIONEL": "e27", "RÍOS ENZO LIONEL": "e27",
    "ROMA": "e28", "JARA ROMA": "e28", "JARA ROMA LAUNEL": "e28",
    "DAN": "e29", "SIEBENHAAR RODRIGUEZ DAN": "e29"
}

def clean_name(name):
    if not name: return ""
    name = str(name).strip().upper()
    name = name.replace("\t", " ").replace("\n", " ")
    name = re.sub(r'\s+', ' ', name)
    return name

def find_docente_id(cell_value):
    val = clean_name(cell_value)
    if not val: return None
    # Look for exact or substring matches in mappings
    for k, v in DOCENTES_MAPPING.items():
        if k in val or val in k:
            return v
    return None

def find_estudiante_id(cell_value):
    val = clean_name(cell_value)
    if not val: return None
    if "LICENCIA" in val or "PARO" in val:
        return "LICENCIA"
    # Try exact match or substring in mapping keys
    for k, v in STUDENTS_MAPPING.items():
        k_clean = k.upper()
        if k_clean in val or val in k_clean:
            return v
    return None

def format_time(t):
    if not t: return None
    if isinstance(t, str):
        # e.g. "13:20" or "1320"
        t = t.replace(":", "").strip()
        if len(t) == 4:
            return f"{t[0:2]}:{t[2:4]}"
        return t
    # datetime.time
    return t.strftime("%H:%M")

def parse_sheet_ini_pri_esp(sheet):
    asignaciones = []
    current_docente_id = None
    
    # Days are Lunes, Martes, Miércoles, Jueves, Viernes
    # Cols:
    # Lunes: Col 2 (student), Col 3 (ing), Col 4 (eg)
    # Martes: Col 5 (student), Col 6 (ing), Col 7 (eg)
    # Miércoles: Col 8 (student), Col 9 (ing), Col 10 (eg)
    # Jueves: Col 11 (student), Col 12 (ing), Col 13 (eg)
    # Viernes: Col 14 (student), Col 15 (ing), Col 16 (eg)
    
    dias_cols = {
        "Lunes": (2, 3, 4),
        "Martes": (5, 6, 7),
        "Miércoles": (8, 9, 10),
        "Jueves": (11, 12, 13),
        "Viernes": (14, 15, 16)
    }

    # Start from row 2
    for r in range(2, sheet.max_row + 1):
        cell_doc = sheet.cell(r, 1).value
        if cell_doc:
            doc_id = find_docente_id(cell_doc)
            if doc_id:
                current_docente_id = doc_id
        
        if not current_docente_id:
            continue
            
        for dia, (sc, ic, ec) in dias_cols.items():
            est_val = sheet.cell(r, sc).value
            if est_val:
                est_id = find_estudiante_id(est_val)
                if est_id:
                    hinicio = format_time(sheet.cell(r, ic).value)
                    hfin = format_time(sheet.cell(r, ec).value)
                    
                    # If time is missing, check if it's Licencia
                    if est_id == "LICENCIA":
                        continue # Licencias are handled separately, or we skip
                    
                    if not hinicio or not hfin:
                        # Fallback default class time
                        hinicio = "13:20"
                        hfin = "14:30"
                        
                    asig_id = f"as_{current_docente_id}_{dia}_{r}"
                    asignaciones.append({
                        "id": asig_id,
                        "docenteId": current_docente_id,
                        "estudianteId": est_id,
                        "dia": dia,
                        "horaInicio": hinicio,
                        "horaFin": hfin
                    })
    return asignaciones

def parse_sheet_secundaria(sheet):
    asignaciones = []
    current_docente_id = None
    
    # Days are Lunes, Martes, Miércoles, Jueves, Viernes
    # Cols:
    # Lunes: Col 2 (student), Col 3 (ing), Col 4 (eg)
    # Martes: Col 5 (student), Col 6 (ing), Col 7 (eg)
    # Miércoles: Col 8 (student), Col 9 (ing), Col 10 (eg)
    # Jueves: Col 11 (student), Col 12 (ing), Col 13 (eg)
    # Viernes: Col 14 (student), Col 15 (ing), Col 16 (eg)
    
    dias_cols = {
        "Lunes": (2, 3, 4),
        "Martes": (5, 6, 7),
        "Miércoles": (8, 9, 10),
        "Jueves": (11, 12, 13),
        "Viernes": (14, 15, 16)
    }

    # Start from row 2
    for r in range(2, sheet.max_row + 1):
        cell_doc = sheet.cell(r, 1).value
        if cell_doc:
            doc_id = find_docente_id(cell_doc)
            if doc_id:
                current_docente_id = doc_id
        
        if not current_docente_id:
            continue
            
        for dia, (sc, ic, ec) in dias_cols.items():
            est_val = sheet.cell(r, sc).value
            if est_val:
                est_id = find_estudiante_id(est_val)
                if est_id:
                    hinicio = format_time(sheet.cell(r, ic).value)
                    hfin = format_time(sheet.cell(r, ec).value)
                    
                    if est_id == "LICENCIA":
                        continue
                    
                    # POT classes have empty times but are active
                    if "T. P." in clean_name(est_val) or "POT" in clean_name(est_val) or "BS AS" in clean_name(est_val):
                        # These are POT sessions or similar, let's keep them as normal with standard times
                        pass
                    
                    if not hinicio or not hfin:
                        # Secondary usually starts at 13:20 or 14:00
                        hinicio = "13:20"
                        hfin = "14:30"
                        
                    asig_id = f"as_{current_docente_id}_{dia}_{r}"
                    asignaciones.append({
                        "id": asig_id,
                        "docenteId": current_docente_id,
                        "estudianteId": est_id,
                        "dia": dia,
                        "horaInicio": hinicio,
                        "horaFin": hfin
                    })
    return asignaciones

def main():
    path = 'Horarios_Mayo/1805_2026/SPOT_GENERAL_-18 al 22 de MAYO-2026.xlsx'
    wb = openpyxl.load_workbook(path, data_only=True)
    
    asigs_pri = parse_sheet_ini_pri_esp(wb['INICIAL - PRIMARIA - ESPECIAL'])
    asigs_sec = parse_sheet_secundaria(wb['SECUNDARIA'])
    
    all_asigs = asigs_pri + asigs_sec
    
    print(f"Parsed {len(asigs_pri)} primary assignments and {len(asigs_sec)} secondary assignments. Total: {len(all_asigs)}")
    
    # Generate new js/data.js with new version
    # Let's read the current js/data.js and replace the asignaciones section, and bump local storage key to v30
    with open('js/data.js', 'r', encoding='utf-8') as f:
        content = f.read()
        
    # Bump localStorage key
    content = content.replace("escuela302_data_v29", "escuela302_data_v30")
    # Bump semana title in CONFIG
    content = re.sub(r'semana:\s*"[^"]*"', 'semana: "18 al 22 de Mayo de 2026"', content)
    
    # Replace asignaciones: [ ... ]
    # Locate DEFAULT_DATA and replace asignaciones
    import json
    asigs_js = json.dumps(all_asigs, indent=4, ensure_ascii=False)
    
    # Format asignaciones to look like JS
    asigs_js = asigs_js.replace('"id":', 'id:').replace('"docenteId":', 'docenteId:').replace('"estudianteId":', 'estudianteId:').replace('"dia":', 'dia:').replace('"horaInicio":', 'horaInicio:').replace('"horaFin":', 'horaFin:')
    
    # Strip outer brackets to avoid double nesting
    asigs_js = asigs_js.strip()
    if asigs_js.startswith('[') and asigs_js.endswith(']'):
        asigs_js = asigs_js[1:-1].strip()
    
    # Locate asignaciones: [ ... ] in DEFAULT_DATA
    pattern = r'(asignaciones:\s*\[)(.*?)(\],\s*novedades:)'
    match = re.search(pattern, content, re.DOTALL)
    if match:
        new_content = content[:match.start(2)] + "\n  " + asigs_js + "\n  " + content[match.end(2):]
        with open('js/data.js', 'w', encoding='utf-8') as fw:
            fw.write(new_content)
        print("js/data.js updated successfully with new week schedules!")
    else:
        print("Could not locate asignaciones array inside DEFAULT_DATA in js/data.js!")

if __name__ == '__main__':
    main()
